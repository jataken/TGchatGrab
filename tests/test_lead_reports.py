"""С8: цифры отчёта по воронке и источникам сходятся с содержимым
lead_events/bot_leads, экспорт в Excel строит все четыре листа, и тот же
пресет запускается через уже существующее расписание выгрузок.
"""
import asyncio
import datetime as dt
import os
import shutil
import sys
import tempfile
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from chatgrab.paths import Paths
from chatgrab.db.database import Database
from chatgrab.core import lead as lead_domain
from chatgrab.core import lead_report
from chatgrab.services.export_service import ExportParams, ExportService
from chatgrab.services.export_schedule_service import ExportScheduleService

base = os.path.join(tempfile.gettempdir(), "cgleadreports")
shutil.rmtree(base, ignore_errors=True)
paths = Paths(Path(base))
paths.ensure()
db = Database(paths.db_path)

print("== core.lead_report.conversion(): чистая арифметика ==")
assert lead_report.conversion(0, 0, 0) == {
    "total": 0, "won": 0, "lost": 0, "in_progress": 0, "conversion_pct": 0.0}
assert lead_report.conversion(10, 4, 2) == {
    "total": 10, "won": 4, "lost": 2, "in_progress": 4, "conversion_pct": 40.0}
# won+lost > total не должно уходить в отрицательный in_progress —
# такого на практике не бывает (won/lost взаимоисключающие статусы одного
# лида), но функция не должна врать при некорректном вводе.
assert lead_report.conversion(1, 1, 1)["in_progress"] == 0
print("  ok")

print("\n== готовим известный набор заявок ==")
db.add_chat(501, "Чат А", "chatA", "all", None)
db.add_chat(502, "Чат Б", "chatB", "all", None)
dir1 = db.add_direction("Косметическое сырьё")
dir2 = db.add_direction("Батарейки")

now = dt.datetime.now().astimezone()
date_from = (now - dt.timedelta(hours=1)).isoformat(timespec="seconds")

lead1 = db.add_lead(None, None, {}, status=lead_domain.NEW, display_name="Ирина",
                     source_chat_id=501, direction_id=dir1,
                     source_type=lead_domain.SOURCE_TYPE_CHAT, event_source=lead_domain.EVENT_SOURCE_MANUAL)
lead2 = db.add_lead(None, None, {}, status=lead_domain.NEW, display_name="Пётр",
                     source_chat_id=501, direction_id=dir1,
                     source_type=lead_domain.SOURCE_TYPE_CHAT, event_source=lead_domain.EVENT_SOURCE_MANUAL)
lead3 = db.add_lead(None, None, {}, status=lead_domain.NEW, display_name="Олег",
                     source_chat_id=501, direction_id=dir2,
                     source_type=lead_domain.SOURCE_TYPE_CHAT, event_source=lead_domain.EVENT_SOURCE_MANUAL)
lead4 = db.add_lead(None, None, {}, status=lead_domain.NEW, display_name="Анна",
                     source_chat_id=502, direction_id=dir2,
                     source_type=lead_domain.SOURCE_TYPE_CHAT, event_source=lead_domain.EVENT_SOURCE_MANUAL)
lead5 = db.add_lead(None, None, {}, status=lead_domain.NEW, display_name="Сергей",
                     source_chat_id=502, direction_id=None,
                     source_type=lead_domain.SOURCE_TYPE_CHAT, event_source=lead_domain.EVENT_SOURCE_MANUAL)
lead6 = db.add_lead(None, None, {}, status=lead_domain.NEW, display_name="Мария",
                     source_chat_id=None, direction_id=dir1,
                     source_type=lead_domain.SOURCE_TYPE_MANUAL, event_source=lead_domain.EVENT_SOURCE_MANUAL)
# За пределами периода — не должна попадать ни в одну сводку.
lead7 = db.add_lead(None, None, {}, status=lead_domain.NEW, display_name="Старая",
                     source_chat_id=501, direction_id=dir1,
                     source_type=lead_domain.SOURCE_TYPE_CHAT, event_source=lead_domain.EVENT_SOURCE_MANUAL)
old_ts = (now - dt.timedelta(days=2)).isoformat(timespec="seconds")
db.execute("UPDATE bot_leads SET created_at = ? WHERE id = ?", (old_ts, lead7))
db.set_lead_status(lead7, lead_domain.WON, source=lead_domain.EVENT_SOURCE_MANUAL)

# lead1, lead4 проходят через «отправлено КП» на известном расстоянии от
# created_at, потом закрываются сделкой.
db.set_lead_status(lead1, lead_domain.QUOTE_SENT, source=lead_domain.EVENT_SOURCE_MANUAL)
db.set_lead_status(lead1, lead_domain.WON, source=lead_domain.EVENT_SOURCE_MANUAL)
lead1_created = db.get_lead(lead1)["created_at"]
quote1_at = (dt.datetime.fromisoformat(lead1_created) + dt.timedelta(days=2)).isoformat(timespec="seconds")
db.execute(
    "UPDATE lead_events SET created_at = ? WHERE lead_id = ? AND kind = 'status' AND to_status = ?",
    (quote1_at, lead1, lead_domain.QUOTE_SENT),
)

db.set_lead_status(lead4, lead_domain.QUOTE_SENT, source=lead_domain.EVENT_SOURCE_MANUAL)
db.set_lead_status(lead4, lead_domain.WON, source=lead_domain.EVENT_SOURCE_MANUAL)
lead4_created = db.get_lead(lead4)["created_at"]
quote4_at = (dt.datetime.fromisoformat(lead4_created) + dt.timedelta(days=4)).isoformat(timespec="seconds")
db.execute(
    "UPDATE lead_events SET created_at = ? WHERE lead_id = ? AND kind = 'status' AND to_status = ?",
    (quote4_at, lead4, lead_domain.QUOTE_SENT),
)

db.set_lead_status(lead2, lead_domain.LOST, reject_reason="не устроила цена",
                   source=lead_domain.EVENT_SOURCE_MANUAL)
db.set_lead_status(lead5, lead_domain.LOST, reject_reason="не отвечает",
                   source=lead_domain.EVENT_SOURCE_MANUAL)
db.set_lead_status(lead3, lead_domain.QUALIFIED, source=lead_domain.EVENT_SOURCE_MANUAL)
# lead6 остаётся NEW

print("  ok")

print("\n== конверсия по источникам ==")
by_source = {r["chat_id"]: r for r in db.leads_report_by_source(date_from, None)}
print("  ", {k: dict(v) for k, v in by_source.items()})
assert by_source[501]["total"] == 3 and by_source[501]["won"] == 1 and by_source[501]["lost"] == 1
assert by_source[502]["total"] == 2 and by_source[502]["won"] == 1 and by_source[502]["lost"] == 1
assert by_source[None]["total"] == 1 and by_source[None]["chat_title"] is None
assert sum(r["total"] for r in by_source.values()) == 6, "заявка за пределами периода не должна попасть в отчёт"
print("  ok")

print("\n== конверсия по направлениям ==")
by_direction = {r["direction_id"]: r for r in db.leads_report_by_direction(date_from, None)}
print("  ", {k: dict(v) for k, v in by_direction.items()})
assert by_direction[dir1]["total"] == 3 and by_direction[dir1]["won"] == 1 and by_direction[dir1]["lost"] == 1
assert by_direction[dir2]["total"] == 2 and by_direction[dir2]["won"] == 1 and by_direction[dir2]["lost"] == 0
assert by_direction[None]["total"] == 1 and by_direction[None]["lost"] == 1
print("  ok")

print("\n== средний срок от первого касания до КП ==")
avg_days = db.avg_days_to_quote(date_from, None)
print("  среднее:", avg_days)
assert avg_days is not None and abs(avg_days - 3.0) < 0.01, avg_days
print("  ok")

print("\n== причины отказов сгруппированы ==")
reasons = {r["reject_reason"]: r["c"] for r in db.reject_reasons_report(date_from, None)}
print("  ", reasons)
assert reasons == {"не устроила цена": 1, "не отвечает": 1}
print("  ok")

print("\n== вне периода — пусто, но не падает ==")
future_from = (now + dt.timedelta(days=10)).isoformat(timespec="seconds")
assert db.leads_report_by_source(future_from, None) == [] or \
    all(r["total"] == 0 for r in db.leads_report_by_source(future_from, None))
assert db.avg_days_to_quote(future_from, None) is None
print("  ok")

print("\n== экспорт в Excel: все четыре листа, число лидов сходится ==")
export_service = ExportService(db, paths)
params = ExportParams(chat_ids=[], kind="leads_report", date_from=date_from,
                       folder=str(paths.exports_dir))
result = export_service.run(params)
assert result.row_count == 6, result.row_count
assert len(result.output_paths) == 1
xlsx_path = Path(result.output_paths[0])
assert xlsx_path.exists()

from openpyxl import load_workbook
wb = load_workbook(xlsx_path)
assert set(wb.sheetnames) == {"Источники", "Направления", "Сводка", "Причины отказов"}, wb.sheetnames
ws_source = wb["Источники"]
assert ws_source.max_row == 1 + len(by_source), (ws_source.max_row, len(by_source))
ws_reasons = wb["Причины отказов"]
assert ws_reasons.max_row == 1 + len(reasons)
print("  ok —", xlsx_path.name)

print("\n== тот же пресет запускается через существующее расписание выгрузок ==")
db.save_preset("Воронка за 30 дней", {**params.__dict__, "preset_name": "Воронка за 30 дней"})
assert any(p["name"] == "Воронка за 30 дней" for p in db.list_presets())

logs = []
sched = ExportScheduleService(db, export_service, on_log=lambda t, tone="": logs.append((tone, t)))
sid = db.add_export_schedule("Воронка за 30 дней", every_hours=168, at_hour=0)
n = asyncio.run(sched.tick(now=dt.datetime(2026, 8, 14, 12, 0)))
row = [r for r in db.list_export_schedules() if r["id"] == sid][0]
print("  запусков:", n, "результат:", row["last_result"])
assert n == 1
assert "лид" in row["last_result"], "для leads_report в результате должно быть про лидов, не про сообщения"
assert logs and logs[-1][0] == "ok"
print("  ok")

db.close()
print("\nТЕСТ ПРОЙДЕН: отчёт по воронке считается по lead_events и выгружается через расписание")
