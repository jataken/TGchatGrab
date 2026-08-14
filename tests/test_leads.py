"""Репозиторий лидов: создание с историей, валидированная смена статуса,
заметки, переписка по отправителю.
"""
import os, sys
import json
import shutil
import tempfile
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from chatgrab.paths import Paths
from chatgrab.db.database import Database
from chatgrab.core import lead as lead_domain

base = os.path.join(tempfile.gettempdir(), "cgleads")
shutil.rmtree(base, ignore_errors=True)
paths = Paths(Path(base))
paths.ensure()
db = Database(paths.db_path)

db.add_chat(1001, "Косметическое сырьё · Биржа", "cosmo", "all", None)
bot_id = db.add_bot("Приёмная", "userbot", None, "custom", "@manager")
contact_id = db.upsert_contact(555, "irina")
direction_id = db.add_direction("Косметическое сырьё", keywords=["глицерин"])

print("== создание пишет и лид, и первую запись истории ==")
lead_id = db.add_lead(
    contact_id, bot_id, {"что": "глицерин"}, status="new",
    username="irina", display_name="Ирина", phone="+7 921 000-00-00",
    source_chat_id=1001, source_type=lead_domain.SOURCE_TYPE_CHAT,
    direction_id=direction_id, product="глицерин", volume="2 тонны",
    event_source=lead_domain.EVENT_SOURCE_RULE,
)
lead = db.get_lead(lead_id)
print("  ", dict(lead))
assert lead["status"] == "new"
assert lead["owner"] == lead_domain.DEFAULT_OWNER
assert lead["source_type"] == lead_domain.SOURCE_TYPE_CHAT
assert lead["direction_id"] == direction_id
assert lead["attachments"] == "[]"

events = db.list_lead_events(lead_id)
print("  история:", [dict(e) for e in events])
assert len(events) == 1
assert events[0]["kind"] == "created"
assert events[0]["source"] == "rule"

print("\n== старые вызовы add_lead (позиционные, без новых полей) не ломаются ==")
old_style_id = db.add_lead(contact_id, bot_id, {"text": "просто текст"}, status="new")
old_lead = db.get_lead(old_style_id)
assert old_lead["source_type"] == "bot", "значение по умолчанию для старых вызовов"
assert old_lead["owner"] == lead_domain.DEFAULT_OWNER
print("  ok")

print("\n== смена статуса пишет и поле, и историю ==")
db.set_lead_status(lead_id, lead_domain.QUALIFIED, source=lead_domain.EVENT_SOURCE_MANUAL)
lead = db.get_lead(lead_id)
assert lead["status"] == lead_domain.QUALIFIED
events = db.list_lead_events(lead_id)
print("  ", [(e["kind"], e["from_status"], e["to_status"]) for e in events])
assert events[-1]["kind"] == "status"
assert events[-1]["from_status"] == "new"
assert events[-1]["to_status"] == lead_domain.QUALIFIED

print("\n== «отказ» без причины отклоняется, ничего не меняется ==")
before_status = db.get_lead(lead_id)["status"]
before_events = len(db.list_lead_events(lead_id))
try:
    db.set_lead_status(lead_id, lead_domain.LOST)
    raised = False
except ValueError as e:
    raised = True
    print("  отклонено:", e)
assert raised, "переход в LOST без причины должен быть отклонён"
assert db.get_lead(lead_id)["status"] == before_status, "статус не должен был измениться"
assert len(db.list_lead_events(lead_id)) == before_events, "лишняя запись истории не должна появиться"

print("\n== «отказ» с причиной проходит и сохраняет её ==")
db.set_lead_status(lead_id, lead_domain.LOST, reject_reason="не устроила цена")
lead = db.get_lead(lead_id)
assert lead["status"] == lead_domain.LOST
assert lead["reject_reason"] == "не устроила цена"
print("  ok")

print("\n== неизвестный статус отклоняется ==")
try:
    db.set_lead_status(lead_id, "выдуманный")
    raised = False
except ValueError:
    raised = True
assert raised
print("  ok")

print("\n== заметки добавляются в историю, пустые — нет ==")
before = len(db.list_lead_events(lead_id))
db.add_lead_note(lead_id, "  ")
assert len(db.list_lead_events(lead_id)) == before, "пустая заметка не должна писаться"
db.add_lead_note(lead_id, "Перезвонить через неделю")
events = db.list_lead_events(lead_id)
assert events[-1]["kind"] == "note"
assert events[-1]["text"] == "Перезвонить через неделю"
print("  ok,", len(events), "записей истории всего")

print("\n== история идёт по порядку времени ==")
kinds = [e["kind"] for e in db.list_lead_events(lead_id)]
print("  ", kinds)
assert kinds == ["created", "status", "status", "note"]

print("\n== вложения через set_lead_field (список -> JSON) ==")
db.set_lead_field(lead_id, attachments=["quotes/КП_Аврора.pdf"])
lead = db.get_lead(lead_id)
assert json.loads(lead["attachments"]) == ["quotes/КП_Аврора.pdf"]
print("  ok")

print("\n== переписка по отправителю ==")
for i in range(1, 4):
    db.upsert_message({"chat_id": 1001, "message_id": i, "chat_title": "Биржа",
        "date": f"2026-08-{i:02d}T10:00:00", "edited_date": None,
        "sender_id": 555, "sender_username": "irina", "sender_display_name": "Ирина",
        "text": f"сообщение {i}", "reply_to_message_id": None, "forwarded_from": None,
        "media_type": None, "media_caption": None, "media_path": None,
        "views": None, "link": "", "is_hidden": 0, "char_len": 0,
        "is_reply": 0, "is_forward": 0})
db.upsert_message({"chat_id": 1001, "message_id": 99, "chat_title": "Биржа",
    "date": "2026-08-05T10:00:00", "edited_date": None,
    "sender_id": 777, "sender_username": "chужой", "sender_display_name": "Чужой",
    "text": "не имеет отношения", "reply_to_message_id": None, "forwarded_from": None,
    "media_type": None, "media_caption": None, "media_path": None,
    "views": None, "link": "", "is_hidden": 0, "char_len": 0,
    "is_reply": 0, "is_forward": 0})
correspondence = db.lead_correspondence(555)
print("  найдено сообщений:", len(correspondence), [m["text"] for m in correspondence])
assert len(correspondence) == 3
assert all(m["sender_id"] == 555 for m in correspondence)
assert correspondence[0]["date"] > correspondence[-1]["date"], "должно идти от новых к старым"

print("\n== воронка сворачивает новую вокабуляр в три старых бакета ==")
# На этом этапе: lead_id в LOST, old_style_id в NEW.
funnel = db.leads_funnel()
print("  ", funnel)
assert funnel["new"] == 1
assert funnel["in_progress"] == 0
assert funnel["closed"] == 1, "LOST должен попадать в «закрыты», а не пропадать"

print("\n== экспорт в Excel подставляет русскую метку, а не сырой код статуса ==")
import tempfile as _tempfile
from chatgrab.bots.export import export_leads_xlsx
from chatgrab.paths import Paths as _Paths
export_dir = os.path.join(base, "exports")
xlsx_path = export_leads_xlsx(db, _Paths(Path(base)), folder=export_dir)
from openpyxl import load_workbook
wb = load_workbook(xlsx_path)
ws = wb.active
status_cells = [row[4].value for row in ws.iter_rows(min_row=2)]
print("  статусы в выгрузке:", status_cells)
assert "отказ" in status_cells, "должна быть русская метка, а не код 'lost'"
assert "lost" not in status_cells and "qualified" not in status_cells

db.close()
print("\nТЕСТ ПРОЙДЕН: репозиторий лидов работает")
