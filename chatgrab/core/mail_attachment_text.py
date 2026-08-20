"""П3: attachment bytes -> plain text, for search and for the reading-text
panel of the attachment viewer. No sqlite, no Qt — pure functions over a
path on disk, testable the same way core/mail_thread.py is.

PDF isn't here: QPdfDocument.getAllText() already does that job, and
QPdfDocument is a Qt object that has to be built on the GUI thread (see
ui/screens/mail/attachment_view.py) — this module only covers formats a
worker thread can parse on its own (docx, xlsx, plain text), the ones
services/mail_service.py's extract_attachment_text() calls off the event
loop the same way it already calls fetch_body().

docx and xlsx are both zip archives — the one thing every extractor here
has to distrust before it distrusts anything else, per PLAN.md's mail
invariant that a message body/attachment is untrusted input. A crafted
archive can declare a tiny file on disk while unpacking to gigabytes, or
pack tens of thousands of members — either would stall or crash the app
handed straight to zipfile/openpyxl, so _check_zip_safe() rejects both
from the central directory, before any member is actually decompressed.
"""
from __future__ import annotations

import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

MAX_ZIP_UNCOMPRESSED = 200 * 1024 * 1024  # 200 MB unpacked, total
MAX_ZIP_MEMBERS = 10_000
MAX_XLSX_SHEETS = 200
MAX_XLSX_ROWS_PER_SHEET = 5_000
MAX_XLSX_COLS_PER_SHEET = 200
MAX_PLAIN_TEXT_BYTES = 5 * 1024 * 1024

_DOCX_EXT = ".docx"
_XLSX_EXT = ".xlsx"
_PLAIN_TEXT_EXTENSIONS = {".txt", ".csv", ".log"}

_W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


class AttachmentParseError(Exception):
    """The file claims to be a format this module handles but isn't a
    valid one — corrupt, truncated, or simply not what its extension
    says. Callers treat this as "stays unindexed", not as a crash."""


class AttachmentTooLargeError(Exception):
    """Rejected by _check_zip_safe() before any real parsing started."""


def _check_zip_safe(path: str) -> None:
    try:
        with zipfile.ZipFile(path) as zf:
            infos = zf.infolist()
            if len(infos) > MAX_ZIP_MEMBERS:
                raise AttachmentTooLargeError(
                    f"в архиве {len(infos)} файлов — больше предела {MAX_ZIP_MEMBERS}")
            total = sum(i.file_size for i in infos)
            if total > MAX_ZIP_UNCOMPRESSED:
                raise AttachmentTooLargeError(
                    f"распакованный размер {total} байт — больше предела {MAX_ZIP_UNCOMPRESSED}")
    except zipfile.BadZipFile as e:
        raise AttachmentParseError(f"файл повреждён или не является zip-архивом: {e}") from e


def extract_docx_text(path: str) -> str:
    """Every paragraph's text, in document order — including paragraphs
    inside table cells, since a OOXML table cell's content is itself made
    of w:p paragraphs. That naturally covers "paragraphs and tables" per
    PLAN.md's П3 wording without separate table-walking code; it doesn't
    preserve column alignment, but the plan is explicit that layout isn't
    the goal here ("вёрстка не воспроизводится и не должна")."""
    _check_zip_safe(path)
    try:
        with zipfile.ZipFile(path) as zf:
            with zf.open("word/document.xml") as f:
                root = ET.parse(f).getroot()
    except (zipfile.BadZipFile, KeyError, ET.ParseError) as e:
        raise AttachmentParseError(f"не удалось разобрать .docx: {e}") from e

    lines = []
    for p in root.iter(f"{_W_NS}p"):
        text = "".join(t.text or "" for t in p.iter(f"{_W_NS}t"))
        if text.strip():
            lines.append(text)
    return "\n".join(lines)


def _load_xlsx_safe(path: str):
    _check_zip_safe(path)
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    # openpyxl's own exceptions for a malformed workbook vary by what's
    # wrong (bad zip, missing part, invalid XML) — caught broadly here
    # because every one of them means the same thing to a caller: this
    # attachment doesn't parse, index nothing, don't crash the sync/view.
    except Exception as e:  # noqa: BLE001 — untrusted external file, see module docstring
        raise AttachmentParseError(f"не удалось разобрать .xlsx: {e}") from e


def extract_xlsx_text(path: str) -> str:
    wb = _load_xlsx_safe(path)
    try:
        parts = []
        for ws in wb.worksheets[:MAX_XLSX_SHEETS]:
            for row in ws.iter_rows(max_row=MAX_XLSX_ROWS_PER_SHEET, values_only=True):
                cells = [str(v) for v in row if v is not None]
                if cells:
                    parts.append(" ".join(cells))
        return "\n".join(parts)
    finally:
        wb.close()


def read_xlsx_grid(path: str) -> list[tuple[str, list[list[str]]]]:
    """Sheet name -> rows of display strings, for the attachment viewer's
    table widget. Same row/size caps as extract_xlsx_text() — a viewer
    asked to render a five-million-row sheet is as much a hang risk as
    indexing one."""
    wb = _load_xlsx_safe(path)
    try:
        sheets = []
        for ws in wb.worksheets[:MAX_XLSX_SHEETS]:
            rows = []
            # max_col fixes the width every row tuple is padded out to —
            # passing the flat cap unconditionally would pad a 3-column
            # sheet out to MAX_XLSX_COLS_PER_SHEET empty cells per row, so
            # it's only ever narrowed from the sheet's own real extent.
            col_cap = min(ws.max_column or 1, MAX_XLSX_COLS_PER_SHEET)
            for row in ws.iter_rows(max_row=MAX_XLSX_ROWS_PER_SHEET,
                                     max_col=col_cap, values_only=True):
                rows.append(["" if v is None else str(v) for v in row])
            sheets.append((ws.title, rows))
        return sheets
    finally:
        wb.close()


def extract_plain_text(path: str) -> str:
    data = Path(path).read_bytes()[:MAX_PLAIN_TEXT_BYTES]
    return data.decode("utf-8", errors="replace")


def read_csv_grid(path: str) -> list[list[str]]:
    """П9: a .csv's rows as a grid, same shape as read_xlsx_grid()'s
    per-sheet rows — core/mail_lead_extract.py's table-field extraction
    works on either without caring which. Same size cap as
    extract_plain_text(): a hostile multi-hundred-MB "csv" attachment is
    as much a hang risk here as anywhere else untrusted input is parsed
    (see this module's own docstring)."""
    import csv
    import io
    data = Path(path).read_bytes()[:MAX_PLAIN_TEXT_BYTES]
    text = data.decode("utf-8", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    return [row for row in reader]


def extract_text_for_search(path: str, filename: str | None, content_type: str | None = None) -> str | None:
    """Dispatch by extension. None means "not a format this module
    handles" (PDF, photos, legacy .doc/.xls) — not an error, just nothing
    for the caller to store. Raises AttachmentParseError/
    AttachmentTooLargeError for a recognized extension that doesn't
    actually parse; the caller (MailService.extract_attachment_text)
    catches both and leaves the attachment unindexed rather than failing
    the sync it ran during."""
    ext = Path(filename or path).suffix.lower()
    if ext == _DOCX_EXT:
        return extract_docx_text(path)
    if ext == _XLSX_EXT:
        return extract_xlsx_text(path)
    if ext in _PLAIN_TEXT_EXTENSIONS:
        return extract_plain_text(path)
    return None
