"""Lead export via the same openpyxl writer style as the parser's
export_service — a plain columned table, no separate export engine."""
from __future__ import annotations

import json
from pathlib import Path

from ..core import lead as lead_domain
from ..db.database import Database, now_iso
from ..paths import Paths
from ..services.xlsx_safety import excel_safe


def export_leads_xlsx(db: Database, paths: Paths, bot_id: int | None = None,
                       folder: str | None = None) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    leads = db.list_leads(bot_id=bot_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"
    headers = ["Дата", "Бот", "Контакт", "Telegram ID", "Статус", "Менеджер", "Содержание"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    wrap = Alignment(vertical="top", wrap_text=True)
    for lead in sorted(leads, key=lambda r: r["created_at"]):
        # Not every lead has a bot or a contact behind it any more (С3:
        # manual and message-based creation) — the lead's own identity
        # fields (С2) are what's authoritative, the joined rows are only
        # a fallback for older, bot-sourced leads that never got them.
        contact = db.get_contact(lead["contact_id"]) if lead["contact_id"] else None
        bot = db.get_bot(lead["bot_id"]) if lead["bot_id"] else None
        handle = lead["display_name"] or \
            (f"@{lead['username']}" if lead["username"] else None) or \
            (f"@{contact['username']}" if contact and contact["username"] else "")
        telegram_id = lead["tg_user_id"] or (contact["telegram_id"] if contact else "")
        source_text = bot["name"] if bot else lead_domain.label_for_source_type(lead["source_type"])
        try:
            content = json.loads(lead["content"])
            # Each value comes straight from a Telegram message (scenario
            # answer or raw text) — excel_safe() per-value, not just on the
            # joined summary, since the leading "field: " prefix that keeps
            # the joined string itself safe today is an implementation
            # detail this shouldn't have to keep relying on.
            summary = "; ".join(f"{k}: {excel_safe(v)}" for k, v in content.items()) if content else ""
        except (json.JSONDecodeError, TypeError):
            summary = ""
        ws.append([
            lead["created_at"], excel_safe(source_text),
            excel_safe(handle), telegram_id,
            lead_domain.label_for_status(lead["status"]), excel_safe(lead["manager"] or ""), summary,
        ])
        row_idx = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).alignment = wrap

    widths = {1: 20, 2: 22, 3: 18, 4: 14, 5: 12, 6: 18, 7: 60}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out_folder = Path(folder or paths.exports_dir)
    out_folder.mkdir(parents=True, exist_ok=True)
    suffix = f"_{bot_id}" if bot_id is not None else ""
    name = f"chatgrab_leads{suffix}_{now_iso()[:10]}.xlsx"
    path = out_folder / name
    wb.save(path)
    return path
