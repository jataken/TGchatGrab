"""Export engine: Excel (.xlsx) / JSONL / Markdown, merged or per-chat,
split by token budget / month / single file, incremental, with an
optional zip of the photos referenced by the selection."""
from __future__ import annotations

import io
import json
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from ..core import lead_report
from ..db.database import Database, now_iso
from ..paths import Paths
from .xlsx_safety import excel_safe

DEFAULT_TOKEN_LIMIT = 180_000
DEFAULT_MD_HEADER = (
    "# Выгрузка сообщений Telegram\n\n"
    "Чаты: {chats}\n"
    "Период: {period}\n"
    "Сообщений: {count}\n\n"
    "Поля: автор и его @ник, дата, текст, вложенное медиа — фото/видео/голосовое/"
    "документ (путь рядом с базой, не встроено в файл), ссылка на исходное "
    "сообщение в Telegram.\n"
    "Файлы лежат в подпапках photos|videos|voice|documents/<chat_id>/<message_id> "
    "рядом с базой данных (или в приложенном zip, если он был запрошен при "
    "выгрузке).\n\n---\n\n"
)

_SLUG_RE = re.compile(r"[^a-zA-Zа-яА-ЯёЁ0-9]+")


def slugify(text: str, max_len: int = 28) -> str:
    s = _SLUG_RE.sub("_", text.strip()).strip("_").lower()
    return (s or "chat")[:max_len]


def estimate_tokens(text: str) -> int:
    """chars/4 approximation — good enough for gauging file counts, no
    extra tokenizer dependency needed."""
    return max(1, len(text) // 4)


def _row_tokens(row) -> int:
    """Token estimate for a row, from the stored char_len when the row was
    fetched without its text (the estimate path), or from the text itself
    otherwise."""
    try:
        length = row["char_len"]
    except (IndexError, KeyError):
        length = None
    if length is None:
        try:
            length = len(row["text"] or "")
        except (IndexError, KeyError):
            length = 0
    return max(1, length // 4)


@dataclass
class ExportParams:
    chat_ids: list[int]
    # С8: a second, unrelated thing this same params/preset/schedule
    # machinery can produce — messages is the export screen's chat
    # export (chat_ids/format/split_mode/... below all apply to it);
    # leads_report is the funnel/source report and only reads
    # date_from/date_to/folder/preset_name, ignoring the rest. Kept as
    # one dataclass rather than a second one so a leads_report preset
    # round-trips through the exact same save_preset/load_preset/
    # export_schedule_service path a messages preset already does —
    # that reuse is the whole point of С8's "через существующее
    # расписание выгрузок".
    kind: str = "messages"        # messages | leads_report
    format: str = "xlsx"          # xlsx | jsonl | markdown
    merge: bool = False
    split_mode: str = "tokens"     # tokens | month | none
    token_limit: int = DEFAULT_TOKEN_LIMIT
    date_from: str | None = None
    date_to: str | None = None
    incremental: bool = False
    zip_photos: bool = False
    include_hidden: bool = False
    unique_only: bool = False
    folder: str = ""
    query: str = ""
    author: str = ""
    photos_only: bool = False
    forwards_only: bool = False
    replies_only: bool = False
    markdown_header: str = ""
    preset_name: str | None = None


@dataclass
class ExportEstimate:
    row_count: int
    token_count: int
    file_count: int
    file_names: list[str] = field(default_factory=list)


@dataclass
class ExportResult:
    output_paths: list[str]
    row_count: int
    export_log_id: int


ROW_COLUMNS = [
    "chat_title", "message_id", "date", "edited_date", "sender_display_name",
    "sender_username", "text", "reply_to_message_id", "forwarded_from",
    "media_type", "media_caption", "media_path", "views", "link",
]

_MEDIA_LABELS = {"photo": "фото", "video": "видео", "voice": "голосовое", "document": "документ"}


def text_with_markers(r) -> str:
    """Post text with a leading reply/forward marker, matching what a
    human reading the export needs to make sense of it out of context —
    e.g. "_ответ на сообщение 91713_ Тера поставляет..."."""
    text = r["text"] or ""
    marks = []
    if r["is_reply"]:
        marks.append(f"ответ на сообщение {r['reply_to_message_id']}")
    if r["is_forward"]:
        marks.append(f"переслано от {r['forwarded_from']}" if r["forwarded_from"] else "переслано")
    if not marks:
        return text
    return f"_{'; '.join(marks)}_ {text}"


def media_marker(r) -> str:
    if not r["media_path"]:
        return ""
    label = _MEDIA_LABELS.get(r["media_type"], "файл")
    return f"[{label}: {r['media_path']}]"


class ExportService:
    def __init__(self, db: Database, paths: Paths):
        self.db = db
        self.paths = paths

    # ---- selection -----------------------------------------------------
    def _selection_kwargs(self, params: ExportParams) -> dict:
        min_id_by_chat = None
        if params.incremental:
            min_id_by_chat = self.db.incremental_baseline(params.chat_ids)
        return dict(
            chat_ids=params.chat_ids, date_from=params.date_from, date_to=params.date_to,
            include_hidden=params.include_hidden, unique_only=params.unique_only,
            query=params.query, author=params.author,
            photos_only=params.photos_only, forwards_only=params.forwards_only,
            replies_only=params.replies_only, min_id_by_chat=min_id_by_chat,
        )

    def _selected_rows(self, params: ExportParams):
        return self.db.export_select(**self._selection_kwargs(params))

    def _selected_meta(self, params: ExportParams):
        """Same selection, without message text — enough to count files and
        tokens, cheap enough to re-run on every change of a checkbox."""
        return self.db.export_select_meta(**self._selection_kwargs(params))

    def _chat_title(self, chat_id: int) -> str:
        chat = self.db.get_chat(chat_id)
        return chat["title"] if chat else f"чат {chat_id}"

    def _group_by_chat(self, rows) -> dict[int, list]:
        groups: dict[int, list] = {}
        for r in rows:
            groups.setdefault(r["chat_id"], []).append(r)
        return groups

    def _base_name(self, params: ExportParams, chat_id: int | None) -> str:
        if params.merge or chat_id is None:
            return "chatgrab_все-чаты"
        return f"chatgrab_{slugify(self._chat_title(chat_id))}"

    def _ext(self, params: ExportParams) -> str:
        return {"xlsx": "xlsx", "jsonl": "jsonl", "markdown": "md"}[params.format]

    # ---- chunking --------------------------------------------------------
    def _chunk_rows(self, rows: list, params: ExportParams) -> list[tuple[str, list]]:
        """Return [(label, rows), ...] where label is used in the filename."""
        if not rows:
            return []
        if params.split_mode == "none":
            return [("", rows)]
        if params.split_mode == "month":
            chunks: dict[str, list] = {}
            for r in rows:
                month = str(r["date"])[:7]
                chunks.setdefault(month, []).append(r)
            return sorted(chunks.items())
        # tokens
        out: list[tuple[str, list]] = []
        current: list = []
        current_tokens = 0
        for r in rows:
            t = _row_tokens(r)
            if current and current_tokens + t > params.token_limit:
                out.append(current)
                current, current_tokens = [], 0
            current.append(r)
            current_tokens += t
        if current:
            out.append(current)
        total = len(out)
        return [(f"часть-{i + 1}-из-{total}" if total > 1 else "", chunk)
                for i, chunk in enumerate(out)]

    # ---- estimate (no writes) -------------------------------------------
    def estimate(self, params: ExportParams) -> ExportEstimate:
        rows = self._selected_meta(params)
        token_total = sum(_row_tokens(r) for r in rows)
        names = self._plan_filenames(rows, params)
        return ExportEstimate(row_count=len(rows), token_count=token_total,
                               file_count=len(names), file_names=names)

    def _plan_filenames(self, rows: list, params: ExportParams) -> list[str]:
        names: list[str] = []
        ext = self._ext(params)
        if params.merge:
            for label, _ in self._chunk_rows(rows, params):
                base = self._base_name(params, None)
                names.append(f"{base}{'_' + label if label else ''}.{ext}")
        else:
            for chat_id, chat_rows in self._group_by_chat(rows).items():
                base = self._base_name(params, chat_id)
                for label, _ in self._chunk_rows(chat_rows, params):
                    names.append(f"{base}{'_' + label if label else ''}.{ext}")
        if params.zip_photos and any(r["media_path"] for r in rows):
            names.append("chatgrab_media.zip")
        return names

    # ---- writers -----------------------------------------------------
    def _write_xlsx(self, path: Path, rows: list) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Сообщения"
        headers = ["Дата и время", "Чат", "Автор", "Ник (@)", "Текст", "Медиа"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        wrap = Alignment(vertical="top", wrap_text=True)
        for r in rows:
            username = f"@{r['sender_username']}" if r["sender_username"] else ""
            ws.append([
                r["date"], excel_safe(r["chat_title"]), excel_safe(r["sender_display_name"] or ""),
                excel_safe(username), excel_safe(text_with_markers(r)), excel_safe(media_marker(r)),
            ])
            row_idx = ws.max_row
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).alignment = wrap
            if r["media_path"]:
                media_cell = ws.cell(row=row_idx, column=6)
                # Relative to the exported file — resolves whether the
                # media files sit on disk as-is next to the export, or the
                # accompanying zip gets extracted into the same folder
                # (it preserves this same photos|videos|voice|documents/
                # <chat_id>/<id> layout).
                media_cell.hyperlink = r["media_path"].replace("\\", "/")
                media_cell.style = "Hyperlink"

        widths = {1: 24, 2: 30, 3: 20, 4: 18, 5: 90, 6: 32}
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(path)

    def _write_jsonl(self, path: Path, rows: list) -> None:
        with path.open("w", encoding="utf-8") as f:
            for r in rows:
                f.write(json.dumps({c: r[c] for c in ROW_COLUMNS}, ensure_ascii=False))
                f.write("\n")

    def _write_markdown(self, path: Path, rows: list, params: ExportParams,
                         chats_label: str) -> None:
        header_tpl = params.markdown_header or DEFAULT_MD_HEADER
        period = f"{params.date_from or 'начало'} — {params.date_to or 'сегодня'}"
        header = header_tpl.format(chats=chats_label, period=period, count=len(rows))
        buf = io.StringIO()
        buf.write(header)
        last_day = None
        for r in rows:
            day = str(r["date"])[:10]
            if day != last_day:
                buf.write(f"\n## {day}\n\n")
                last_day = day
            author = r["sender_display_name"] or "—"
            handle = f" (@{r['sender_username']})" if r["sender_username"] else ""
            buf.write(f"**{author}{handle}** · {r['chat_title']} · {r['date']}\n\n")
            if r["forwarded_from"]:
                buf.write(f"_переслано от {r['forwarded_from']}_\n\n")
            if r["reply_to_message_id"]:
                buf.write(f"_ответ на сообщение {r['reply_to_message_id']}_\n\n")
            buf.write(f"{r['text'] or ''}\n\n")
            if r["media_path"]:
                buf.write(f"{media_marker(r)}\n\n")
            buf.write(f"<{r['link']}>\n\n---\n\n")
        path.write_text(buf.getvalue(), encoding="utf-8")

    def _write_chunk(self, path: Path, rows: list, params: ExportParams, chats_label: str) -> None:
        if params.format == "xlsx":
            self._write_xlsx(path, rows)
        elif params.format == "jsonl":
            self._write_jsonl(path, rows)
        else:
            self._write_markdown(path, rows, params, chats_label)

    # ---- run -----------------------------------------------------------
    def run(self, params: ExportParams) -> ExportResult:
        if params.kind == "leads_report":
            return self._run_leads_report(params)
        rows = self._selected_rows(params)
        folder = Path(params.folder or self.paths.exports_dir)
        folder.mkdir(parents=True, exist_ok=True)
        ext = self._ext(params)
        output_paths: list[str] = []
        chats_label = ", ".join(self._chat_title(c) for c in params.chat_ids)

        if params.merge:
            for label, chunk in self._chunk_rows(rows, params):
                base = self._base_name(params, None)
                name = f"{base}{'_' + label if label else ''}.{ext}"
                path = folder / name
                self._write_chunk(path, chunk, params, chats_label)
                output_paths.append(str(path))
        else:
            for chat_id, chat_rows in self._group_by_chat(rows).items():
                base = self._base_name(params, chat_id)
                for label, chunk in self._chunk_rows(chat_rows, params):
                    name = f"{base}{'_' + label if label else ''}.{ext}"
                    path = folder / name
                    self._write_chunk(path, chunk, params, self._chat_title(chat_id))
                    output_paths.append(str(path))

        if params.zip_photos:
            media_rows = [r for r in rows if r["media_path"]]
            if media_rows:
                zip_path = folder / "chatgrab_media.zip"
                with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                    for r in media_rows:
                        src = self.paths.data_dir / r["media_path"]
                        if src.exists():
                            zf.write(src, arcname=r["media_path"])
                output_paths.append(str(zip_path))

        max_id_by_chat: dict[str, int] = {}
        for r in rows:
            cid = str(r["chat_id"])
            max_id_by_chat[cid] = max(max_id_by_chat.get(cid, 0), r["message_id"])

        log_id = self.db.add_export_log(
            created_at=now_iso(),
            chat_ids=json.dumps(params.chat_ids),
            format=params.format,
            date_from=params.date_from,
            date_to=params.date_to,
            merge=1 if params.merge else 0,
            split_mode=params.split_mode,
            token_limit=params.token_limit,
            incremental=1 if params.incremental else 0,
            zip_photos=1 if params.zip_photos else 0,
            include_hidden=1 if params.include_hidden else 0,
            max_message_id_by_chat=json.dumps(max_id_by_chat),
            output_paths=json.dumps(output_paths),
            preset_name=params.preset_name,
        )
        return ExportResult(output_paths=output_paths, row_count=len(rows), export_log_id=log_id)

    # ---- leads report (С8) ------------------------------------------------
    # A different question than the rest of this file — not "which
    # messages", but "how did the leads that came in convert" — sharing
    # the same ExportParams/preset/schedule plumbing rather than a
    # parallel one, per params.kind's own docstring above.
    def _run_leads_report(self, params: ExportParams) -> ExportResult:
        folder = Path(params.folder or self.paths.exports_dir)
        folder.mkdir(parents=True, exist_ok=True)

        by_source = self.db.leads_report_by_source(params.date_from, params.date_to)
        by_direction = self.db.leads_report_by_direction(params.date_from, params.date_to)
        avg_days = self.db.avg_days_to_quote(params.date_from, params.date_to)
        reject_reasons = self.db.reject_reasons_report(params.date_from, params.date_to)

        name = f"chatgrab_отчёт-по-воронке_{now_iso()[:10]}.xlsx"
        path = folder / name
        self._write_leads_report_xlsx(path, by_source, by_direction, avg_days, reject_reasons)

        # by_source's NULL-chat bucket included, so this counts every lead
        # in range exactly once — same total leads_report_by_direction
        # would give, just grouped differently.
        total_leads = sum(r["total"] for r in by_source)
        log_id = self.db.add_export_log(
            created_at=now_iso(), chat_ids=json.dumps([]), format="xlsx",
            date_from=params.date_from, date_to=params.date_to,
            merge=1 if params.merge else 0, split_mode=params.split_mode,
            token_limit=params.token_limit, incremental=0, zip_photos=0,
            include_hidden=0, max_message_id_by_chat=json.dumps({}),
            output_paths=json.dumps([str(path)]), preset_name=params.preset_name,
        )
        return ExportResult(output_paths=[str(path)], row_count=total_leads, export_log_id=log_id)

    def _write_leads_report_xlsx(self, path: Path, by_source: list, by_direction: list,
                                  avg_days: float | None, reject_reasons: list) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws_source = wb.active
        ws_source.title = "Источники"
        self._write_conversion_sheet(ws_source, by_source, "chat_title", "без чата / от бота")

        ws_direction = wb.create_sheet("Направления")
        self._write_conversion_sheet(ws_direction, by_direction, "direction_name", "без направления")

        ws_summary = wb.create_sheet("Сводка")
        ws_summary.append([
            "Средний срок от первого касания до КП, дней",
            round(avg_days, 1) if avg_days is not None else "нет данных",
        ])
        ws_summary["A1"].font = Font(bold=True)
        ws_summary.column_dimensions["A"].width = 48

        ws_reasons = wb.create_sheet("Причины отказов")
        ws_reasons.append(["Причина", "Количество"])
        for cell in ws_reasons[1]:
            cell.font = Font(bold=True)
        for row in reject_reasons:
            ws_reasons.append([excel_safe(row["reject_reason"] or "не указана"), row["c"]])
        ws_reasons.column_dimensions["A"].width = 32

        wb.save(path)

    def _write_conversion_sheet(self, ws, rows: list, label_key: str, empty_label: str) -> None:
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        ws.append(["Название", "Всего", "Сделки", "Отказы", "В работе", "Конверсия, %"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            conv = lead_report.conversion(row["total"], row["won"], row["lost"])
            label = row[label_key] or empty_label
            ws.append([
                excel_safe(label), conv["total"], conv["won"], conv["lost"],
                conv["in_progress"], conv["conversion_pct"],
            ])
        widths = {1: 32, 2: 10, 3: 10, 4: 10, 5: 10, 6: 14}
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # ---- presets -----------------------------------------------------
    def save_preset(self, name: str, params: ExportParams) -> None:
        self.db.save_preset(name, params.__dict__)

    def load_preset(self, name: str) -> ExportParams | None:
        for row in self.db.list_presets():
            if row["name"] == name:
                data = json.loads(row["params"])
                return ExportParams(**data)
        return None
